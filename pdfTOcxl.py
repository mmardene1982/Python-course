
import PyPDF2
import openpyxl
# creating an object
pdfFileObj= open('bha1.pdf', 'rb')
pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
pageObj = pdfReader.getPage(0)
pageline=pageObj.extractText()
lst=pageline.split()

#print(lst)

matchingTools=["Cablehead","Swivel","Bluespark","RSS","PTA"'DECT001',
               'MIT SRO', 'PTA MAPS', 'PTA LWT, MAPS', 'Spear Cement tag', 'Explosive cutter', 'Blue Spark', 'Prime w Halliburton Plug w drif', 'Plug HTHP RSS']
matching=[s for s in matchingTools if any (s in a for a in lst)]

print(matching)
theFile = openpyxl.load_workbook('risk.xlsx')
allSheetNames = theFile.sheetnames

print("All sheet names {} " .format(theFile.sheetnames))